import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# 定義路徑
base_path = Path(__file__).resolve().parents[1] / "data"
base_path.mkdir(parents=True, exist_ok=True)
file_path = base_path / "HK_Birdnest_Store_Visit_Log.xlsx"

# 1. 建立 Visit_Data Sheet (供資料登錄)
headers = [
    "拜訪日期", "店名", "地區", "店鋪類型", "拿貨來源",
    "採購模式 (賣斷/寄賣)", "結款帳期 (天)", "我方報價", "市場反應",
    "產品偏好 (備註)", "老闆個性/風險評估", "綜合評分 (1-5)", "Next Action"
]

data = [
    ["2026-03-07", "範例店鋪 (請刪除)", "上環", "海味店", "大拆商", "賣斷", 30, 7.6, "合理", "大盞/偏白", "老闆豪爽但壓價", 4, "下週帶樣品"]
]

df = pd.DataFrame(data, columns=headers)

# 2. 使用 openpyxl 進行格式美化
wb = Workbook()
ws = wb.active
ws.title = "Visit_Data"

# 寫入標頭
for col_num, column_title in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = column_title
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="BF8B3D", end_color="BF8B3D", fill_type="solid") # 金色
    cell.alignment = Alignment(horizontal="center")

# 寫入資料
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num).value = cell_value

# 調整欄寬
column_widths = [15, 25, 10, 15, 15, 20, 15, 10, 15, 20, 25, 12, 20]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64+i)].width = width

# 3. 建立 Checklist_Printable Sheet (列印專用格式)
ws2 = wb.create_sheet("Checklist_Printable")
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 60

rows = [
    ["🇻🇳 Vietnam Birdnest", "香港門店拜訪清單 (Printable)"],
    ["拜訪日期", "____ / ____ / ____"],
    ["拜訪人員", "________________"],
    ["", ""],
    ["📍 基本資訊", ""],
    ["店名", "__________________________________________________"],
    ["地址", "__________________________________________________"],
    ["聯絡人", "________________"],
    ["", ""],
    ["💵 核心數據 (一間店發一個語音)", ""],
    ["店鋪類型", "[ ] 海味店  [ ] 藥房  [ ] 精品店"],
    ["拿貨來源", "[ ] 大拆商  [ ] 自行進口  [ ] 不清楚"],
    ["採購模式", "[ ] 賣斷 (Cash)  [ ] 寄賣 (Consignment)"],
    ["結款週期", "[ ] 現金 (COD)  [ ] 月結 30  [ ] 月結 60+"],
    ["", ""],
    ["📦 產品需求", ""],
    ["我方報價", "$________ / g"],
    ["市場反應", "[ ] 偏貴  [ ] 合理  [ ] 具競爭力"],
    ["老闆反饋", "__________________________________________________"],
    ["", ""],
    ["⭐ 綜合評價", "[ ] 1星  [ ] 2星  [ ] 3星  [ ] 4星  [ ] 5星 (必攻)"],
]

for r_idx, (label, val) in enumerate(rows, 1):
    ws2.cell(row=r_idx, column=1).value = label
    ws2.cell(row=r_idx, column=2).value = val
    if r_idx == 1:
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws2.cell(row=1, column=1).font = Font(bold=True, size=16)

    if "📍" in label or "💵" in label or "📦" in label or "⭐" in label or "🇻🇳" in label:
        ws2.cell(row=r_idx, column=1).font = Font(bold=True, size=12)
        ws2.cell(row=r_idx, column=1).fill = PatternFill(start_color="F2E6D9", end_color="F2E6D9", fill_type="solid")

wb.save(str(file_path))
print(f"Excel created at: {file_path}")
